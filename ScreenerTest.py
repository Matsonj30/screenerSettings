from datetime import date
from selenium import webdriver
from openpyxl import load_workbook
import selenium
import openpyxl
from selenium.webdriver.chrome.options import Options



today = date.today()

#This is used to retrieve the webpage/ignore annoying errors this retrieval may have
options = webdriver.ChromeOptions()
options.add_argument('--ignore-ssl-errors')
options.add_argument('--ignore-certificate-errors')
website = webdriver.Chrome(executable_path=r"C:/Users/Jared\AppData/Local/Programs/Python/Python38-32/chromedriver.exe", chrome_options=options)


website.get('https://finviz.com/screener.ashx?v=111&f=sh_price_u10,sh_relvol_o10&ft=4')



green = website.find_elements_by_class_name('is-green') #green/red always 3 more in y value than ticker? no idea why
red = website.find_elements_by_class_name('is-red')
tickers = website.find_elements_by_class_name('screener-link-primary') #find each ticker (unique through being a link)
numberOfTickers = int(website.find_element_by_xpath('//*[@id="screener-content"]/table/tbody/tr[3]/td/table/tbody/tr/td[1]').text[7:9]) #future use?
Xvalue = int(str(green[0].location)[6:9]) #seems to change?? will only work if first tiker does not have green P/E 


if numberOfTickers > 20:
    numberOfTickers = 20

sheet = load_workbook("D:/Programming/Python/projects/ScreenerUpdates/HighRelativeVolumePennies.xlsx")
sheetAlter = sheet.worksheets[0] #starting point??

#findStart() returns Int
#iterates through the spreadsheet and finds the highest row that is empty 
def findStart(): #works returns 16 as of right now
    startingPoint = 1
    while sheetAlter.cell(row=startingPoint, column=1).value is not None:
        startingPoint += 1
    return startingPoint    

newStart = findStart() #so we start at the fist open cell
for number in range(numberOfTickers): #should be number of tickers
    
    sheetAlter.cell(row=number+newStart, column=1).value = tickers[number].text
    tickerYCoord = str(tickers[number].location)
    if(len(tickerYCoord)) == 20:                #if y coordinate is 4 digits
        tickerYCoord = int(tickerYCoord[15:19]) #uses smaller indeces than the green/red because the x coordinate is 2 digits instead of 3
    else:                                   
        tickerYCoord = int(tickerYCoord[15:18]) #if y coordinate is 3 digits
                      
   
    for greenNumber in range(len(green)):
        greenBoth = str(green[greenNumber].location)
        greenX = int(greenBoth[6:9])

        if(len(greenBoth)) == 21:
            greenY = int(greenBoth[16:20])  #this gets the coords of a specific green value if its coords are 4 digits
        else:
            greenY = int(greenBoth[16:19])  #this gets the coords of a specific green value if its coords are 3 digits
    
        if(greenY - 3 == tickerYCoord) and greenX == Xvalue:
            sheetAlter.cell(row=number+newStart, column=2).value = float(green[greenNumber].text) #price write in
            sheetAlter.cell(row=number+newStart, column=3).value = float(green[greenNumber+1].text[:-1]) #% change write in
            break #terminates innermost loop 

    for redNumber in range(len(red)):
        redBoth = str(red[redNumber].location)
        redX = int(redBoth[6:9])

        if(len(redBoth)) == 21:
            redY = int(redBoth[16:20]) #this gets the coords of a specific red value if its coords are 4 digits
        else:
            redY = int(redBoth[16:19])  #this gets the coords of a specific red value if its coords are 3 digits
        
        if(redY - 3 == tickerYCoord and redX == Xvalue):
            sheetAlter.cell(row=number+newStart, column=2).value = float(red[redNumber].text) * -1 #maybe change this
            sheetAlter.cell(row=number+newStart, column=3).value = float(red[redNumber+1].text[:-1]) 
            break


sheet.save("D:/Programming/Python/projects/ScreenerUpdates/HighRelativeVolumePennies.xlsx") #have to save or edits wont actually be made
website.close()
