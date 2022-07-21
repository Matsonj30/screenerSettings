import requests
import pandas as pd
from openpyxl import load_workbook
from datetime import date
import pyodbc

#BUGS -> Excel can write twice if the finviz screener has less than two pages of results, the second url
#will return the first ticker even if it doesnt belong on the second page


#startPoint()
#returns the right most column in the excel sheet which is blank
#this is where we will start inputting new tickers found for the day
#PARAMETERS: sheetWrite -> the excel sheet we want to find the starting point of
#RETURNS: NONE
def startPoint(sheetWrite):
    rowNumber = 1
    while sheetWrite.cell(row = rowNumber, column = 1).value != None:
        rowNumber += 1
    return rowNumber


#writeToSQL()
#using data retrieved by finvizData() will insert said values into azure SQL database
#PARAMETERS: data -> array containing lists of strings/floats/ints
#RETURNS: NONE
def writeToSQL(data):
    file = open("D:\Programming\SecretKeysandPass\serverTest.txt")
    password = file.readline()
    connect = pyodbc.connect('DRIVER={SQL Server};SERVER=jaredsdb.database.windows.net;DATABASE=jaredDateBase;UID=matsonj2013@gmail.com@jaredsdb;PWD='+password) #DATABASE != servername
    cursor = connect.cursor() #THIS WORKS
    names = data[0]
    index = 0 
    #need to esnrue every string has a '' around it so we use.format
    for name in names:   
        index += 1
        query = '''INSERT INTO highVolume(ticker,industry,dateFound,volume,mktCap,priceFound,changeWhenFound)
                VALUES('{ticker}','{industry}','{dateFound}',{volume},'{mktCap}',{priceFound},'{changeWhenFound}')
                '''.format(ticker = data[0][index], industry =data[1][index], dateFound = data[2], volume = data[3][index], mktCap =data[4][index], priceFound=data[5][index], changeWhenFound=data[6][index])
        try:
            cursor.execute(query)
        except pyodbc.IntegrityError:
            continue

    cursor.commit()


#writeToExcel()
#using data retrieved by finvizData() will write to excel sheet
#PARAMETERS: data -> array containing lists of strings/floats/ints
#RETURNS: NONE
def writeToExcel(data):
    sheet = load_workbook("D:/Programming/Repositories/screenerSettings/highVolume.xlsx")
    sheetWrite = sheet.worksheets[0]
    startingPoint = startPoint(sheetWrite)
    index = 1 #this index starts at 0 because the pandas dataset does not have a 0th row
    names = data[1]
    
    #this can be done in less lines but its tricky
    for name in names:
        sheetWrite.cell(row = startingPoint, column = 1).value = data[0][index] ##iterate each thing here
        sheetWrite.cell(row = startingPoint, column = 2).value = data[1][index] ##iterate each thing here
        sheetWrite.cell(row = startingPoint, column = 3).value = date.today() ##iterate each thing here
        sheetWrite.cell(row = startingPoint, column = 4).value = data[3][index] ##iterate each thing here
        sheetWrite.cell(row = startingPoint, column = 5).value = data[4][index] ##iterate each thing here
        sheetWrite.cell(row = startingPoint, column = 6).value = data[5][index] ##iterate each thing here
        sheetWrite.cell(row = startingPoint, column = 7).value = data[6][index] ##iterate each thing here
        index += 1
        startingPoint += 1
    sheet.save(("D:/Programming/Repositories/screenerSettings/highVolume.xlsx"))

#finvizData()
#retrieves finviz screener data by utilizing pandas to read html tables on finviz.com
#after gaining the data we want, we put it in an array to pass to writeToSQL() and writeToExcel()
#so we can store the same data in two places as we learn how to work with each better
#PARAMETERS: NONE
#RETURNS: NONE
def finvizData():
    headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
    urls = ['https://finviz.com/screener.ashx?v=111&f=sh_float_u100,sh_relvol_o10&ft=4&o=volume',"https://finviz.com/screener.ashx?v=111&f=sh_float_u100,sh_relvol_o10&ft=4&o=volume&r=21"]
    #need to put something here to stop potential double writes ^^^
    
    #pandas indexing columns and rows we want
    for url in urls:
        screenerPage = requests.get(url, headers = headers).text
        tables = pd.read_html(screenerPage)
        tables = tables[-2] #this is the table we want from the many tables pandas found
        names = tables.iloc[1:, 1] #I think _: is better than 1: even if do the same thing
        industry = tables.iloc[1:, 4]
        marketCap = tables.iloc[1:,6]
        price = tables.iloc[1:,8]
        change = tables.iloc[1:,9]
        volume = tables.iloc[1:,10] #[row selection, column selection] BY NUMBER labelled in the pandas table

        dataToAdd = [names, industry, date.today(), volume, marketCap, price, change]
        writeToExcel(dataToAdd)
        writeToSQL(dataToAdd)

