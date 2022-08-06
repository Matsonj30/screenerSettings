from openpyxl import load_workbook
import pyodbc
from yahoo_fin.stock_info import *
from datetime import date
from readScreener import finvizData


#writeToExcel()
#will grab live price data of tickers that were previously scanned by finvizData() and put them into the Azure DB jaredsDB
#PARAMETERS: NONE
#RETURNS: NONE
def sqlUpdate():
    file = open("D:\Programming\SecretKeysandPass\serverTest.txt")
    password = file.readline()
    connect = pyodbc.connect('DRIVER={SQL Server};SERVER=jaredsdb.database.windows.net;DATABASE=jaredDateBase;UID=matsonj2013@gmail.com@jaredsdb;PWD='+password) #DATABASE != servername
    cursor = connect.cursor() 

    query = '''SELECT ticker, day1, day2, day3, day4, day5, day6, day7, day8, day9, day10, day11, day12, day13, day14 FROM highvolume
            '''
    cursor.execute(query)
    for row in cursor.fetchall():
        for i in range(15):
            if(row[i] == None):
                updateQuery = '''UPDATE highVolume
                                 SET  {dayToUpdate} ='{tickerLiveValue}'
                                 WHERE ticker = '{tickerName}'
                              '''.format(dayToUpdate = 'day'+str(i), tickerLiveValue = round(get_live_price(row[0]),2), tickerName = row[0])
                try:
                    cursor.execute(updateQuery)
                except Exception:
                    print("Ticker does not exist " + row[0])
                break
    cursor.commit()
                                
        
        
        
           


#writeToExcel()
#will grab live price data of tickers that were previously scanned by finvizData() and put them into the excel sheet highVolume.xslx
#PARAMETERS: NONE
#RETURNS: NONE
def excelUpdate():
    sheet = load_workbook("D:/Programming/Repositories/screenerSettings/highVolume.xlsx")
    sheetWrite = sheet.worksheets[0] #starting point??
    index = 2
    while(sheetWrite.cell(row = index, column=1).value != None):
        if(sheetWrite.cell(row = index, column=21).value == None): #checking to see if we need to update this ticker anymore
            for cellXIndex in range(14):
                if(sheetWrite.cell(row = index, column = 8 + cellXIndex).value == None):
                    sheetWrite.cell(row = index, column = 8 + cellXIndex).value = round(get_live_price(sheetWrite.cell(row = index, column=1).value),3) #get live price of company, rounded 3 decimal spots
                    break
        index += 1
    sheet.save("D:/Programming/Repositories/screenerSettings/highVolume.xlsx")
sqlUpdate()
excelUpdate()
finvizData()